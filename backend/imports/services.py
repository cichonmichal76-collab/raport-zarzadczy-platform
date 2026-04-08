import re

from openpyxl import load_workbook

from reports.models import ProductionRecord, RnDRecord

from .models import ImportLog, UploadBatch


VALID_PRODUCTION_STATUSES = {"Gotowe", "W toku", "Stop", "Zaplanowane"}
ORDER_NUMBER_RE = re.compile(r"^\d+/\d+/\d{4}$")


class ImportValidationError(Exception):
    pass


def _sheet_rows(sheet):
    return list(sheet.iter_rows(values_only=True))


def _normalize_headers(row):
    return [str(cell or "").strip().lower() for cell in row]


def _find_header_index(headers, patterns):
    for idx, header in enumerate(headers):
        for pattern in patterns:
            if pattern in header:
                return idx
    return -1


def _find_header_row(rows, marker):
    for idx, row in enumerate(rows):
        values = [str(cell or "").strip().lower() for cell in row]
        if any(marker in value for value in values):
            return idx
    return -1


def _log(batch, level, message):
    ImportLog.objects.create(batch=batch, level=level, message=message)


def _safe_int(value, default=0):
    if value in (None, ""):
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip().replace(" ", "")
    try:
        return int(float(text.replace(",", ".")))
    except (TypeError, ValueError):
        return default


def process_batch(batch: UploadBatch):
    workbook = load_workbook(batch.file.path, data_only=True)
    if batch.source == UploadBatch.Source.PRODUCTION:
        return _import_production(batch, workbook)
    if batch.source == UploadBatch.Source.RND:
        return _import_rnd(batch, workbook)
    raise ImportValidationError("Nieobslugiwane zrodlo wsadu.")


def _value(row, index, default=""):
    if index < 0 or index >= len(row):
        return default
    cell = row[index]
    if cell is None:
        return default
    return cell


def _import_production(batch: UploadBatch, workbook):
    sheet = workbook["Zlecenia"] if "Zlecenia" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = _sheet_rows(sheet)
    header_row = _find_header_row(rows, "nr zlecenia")
    if header_row < 0:
        raise ImportValidationError('Nie znaleziono naglowka "Nr zlecenia".')

    headers = _normalize_headers(rows[header_row])
    indexes = {
        "order_number": _find_header_index(headers, ["nr zlecenia"]),
        "status": _find_header_index(headers, ["status"]),
        "product": _find_header_index(headers, ["produkt"]),
        "group": _find_header_index(headers, ["grupa"]),
        "machine": _find_header_index(headers, ["maszyna"]),
        "completed_units": _find_header_index(headers, ["szt. gotowe", "szt gotowe"]),
        "planned_units": _find_header_index(headers, ["szt. wszystkich", "wszystkich"]),
        "work_time": _find_header_index(headers, ["czas pracy"]),
        "norm_time": _find_header_index(headers, ["normatyw"]),
        "workers": _find_header_index(headers, ["pracownicy", "operatorzy"]),
        "current_state": _find_header_index(headers, ["stan obecny", "uwagi", "opis technologiczny", "komentarz"]),
        "problem": _find_header_index(headers, ["problem"]),
        "solution": _find_header_index(headers, ["rozwiazanie", "rozwiązanie"]),
    }

    if indexes["order_number"] < 0 or indexes["product"] < 0:
        raise ImportValidationError("Brakuje wymaganych kolumn dla produkcji.")

    ProductionRecord.objects.filter(period=batch.period).delete()
    imported_count = 0
    skipped_count = 0

    for row in rows[header_row + 1 :]:
        order_number = str(_value(row, indexes["order_number"])).strip()
        status = str(_value(row, indexes["status"])).strip()
        product = str(_value(row, indexes["product"])).strip()

        if not order_number and not status and not product:
            continue

        if not ORDER_NUMBER_RE.match(order_number) or status not in VALID_PRODUCTION_STATUSES or not product:
            skipped_count += 1
            continue

        ProductionRecord.objects.create(
            period=batch.period,
            order_number=order_number,
            status=status,
            product=product,
            product_group=str(_value(row, indexes["group"])).strip(),
            machine=str(_value(row, indexes["machine"])).strip(),
            completed_units=_safe_int(_value(row, indexes["completed_units"], 0)) if indexes["completed_units"] >= 0 else 0,
            planned_units=_safe_int(_value(row, indexes["planned_units"], 0)) if indexes["planned_units"] >= 0 else 0,
            work_time=str(_value(row, indexes["work_time"])).strip(),
            norm_time=str(_value(row, indexes["norm_time"])).strip(),
            workers=str(_value(row, indexes["workers"])).strip(),
            current_state=str(_value(row, indexes["current_state"])).strip(),
            problem=str(_value(row, indexes["problem"])).strip(),
            solution=str(_value(row, indexes["solution"])).strip(),
        )
        imported_count += 1

    if imported_count == 0:
        raise ImportValidationError("Nie znaleziono rekordow produkcyjnych do importu.")

    _log(batch, "info", f"Produkcja: zapisano {imported_count} rekordow.")
    if skipped_count:
        _log(batch, "warning", f"Produkcja: pominieto {skipped_count} wierszy technicznych lub niepoprawnych.")
    return {"records_imported": imported_count, "rows_skipped": skipped_count}


def _section_text(rows, start_label, end_label):
    active = False
    items = []
    for row in rows:
        first = str(row[0] or "").strip().upper()
        if start_label in first:
            active = True
            continue
        if end_label and end_label in first:
            active = False
            continue
        if not active:
            continue
        value = str((row[1] if len(row) > 1 else row[0]) or "").strip()
        if value:
            items.append(value)
    return "\n".join(items)


def _import_rnd(batch: UploadBatch, workbook):
    RnDRecord.objects.filter(period=batch.period).delete()
    imported_count = 0

    for code in ["BR1", "BR2", "BR3", "BR4", "WBTS", "PNC"]:
        if code not in workbook.sheetnames:
            continue
        sheet = workbook[code]
        rows = _sheet_rows(sheet)
        status = ""
        progress = 0
        trl = 5
        for row in rows:
            if str(row[0] or "").strip() == code:
                progress = _safe_int(row[1], 0)
                trl = max(5, min(9, _safe_int(row[2], 5)))
                status = str(row[4] or "").strip()
                break

        problem = _section_text(rows, "SEKCJA 2", "SEKCJA 3")
        solution = _section_text(rows, "SEKCJA 3", "")
        RnDRecord.objects.create(
            period=batch.period,
            code=code,
            name=code,
            status=status or "W toku",
            progress=progress,
            trl_level=trl,
            current_state=f"Zaimportowano dane dla obszaru {code}.",
            problem=problem,
            solution=solution,
        )
        imported_count += 1

    if imported_count == 0:
        raise ImportValidationError("Nie znaleziono zadnych arkuszy B+R do importu.")

    _log(batch, "info", f"B+R: zapisano {imported_count} rekordow.")
    return {"records_imported": imported_count, "rows_skipped": 0}
